from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create a new Workbook
wb = Workbook()
ws = wb.active

# Data to merge
merge_data = [
    (1, 1),  # Merge cells from row 1, column 1
    (3, 1),  # Merge cells from row 2, column 1
]

    
def merge(merge_data):
    # Perform merging
    for row, col in merge_data:
        start_cell = ws.cell(row=row, column=col)
        end_cell = ws.cell(row=row, column=col+1)  # Assuming merging 2 cells horizontally
        ws.merge_cells(start_cell.coordinate + ':' + end_cell.coordinate)
    
merge(merge_data)
# Assigning values to merged cells
ws.cell(row=1, column=1, value='Merged Cells 1')
ws.cell(row=2, column=1, value='Merged Cells 2')
ws.cell(row=3, column=1, value='Merged Cells 3')

# Save the file
wb.save('example.xlsx')
